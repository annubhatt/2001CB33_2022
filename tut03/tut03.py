from math import ceil
#importing openpyxl, workbook and datetime
import openpyxl
from openpyxl import Workbook
from datetime import datetime

starttime = datetime.now()

def get_octant(x, y, z):
    if(x>=0 and y>=0 and z>=0):
        return(1)
    elif(x<0 and y>=0 and z>=0):
        return(2)
    elif(x<0 and y<0 and z>=0):
        return(3)
    elif(x>=0 and y<0 and z>=0):
        return(4)
    elif(x>=0 and y>=0 and z<0):
        return(-1)
    elif(x<0 and y>=0 and z<0):
        return(-2)
    elif(x<0 and y<0 and z<0):
        return(-3)
    else:
        return(-4)

def octant_longest_subsequence_count():
    # processed values (to be used in program)
    Ud, Vd, Wd, Ov=[], [], [], []

    # values(sum)
    u_sum, v_sum, w_sum=0, 0, 0

    # value count
    n=0
    
    try:

        #  xlsx Input file  
        wb_input=openpyxl.load_workbook("input_octant_longest_subsequence.xlsx")
        input_sh=wb_input.active

        # get the number of rows
        nrow=input_sh.max_row

        # n (no of values) = number of rows(nrow) - header row(1)
        n=nrow-1

        for i in range(2, nrow+1):
            u_sum+=input_sh.cell(i,2).value
            v_sum+=input_sh.cell(i,3).value
            w_sum+=input_sh.cell(i,4).value

        # finding average values
        u_sum/=n; v_sum/=n; w_sum/=n

        #reporting data upto 9 decimal places
        u_sum=int(u_sum*1000000000); u_sum/=1000000000 #typecasting into integer
        v_sum=int(v_sum*1000000000); v_sum/=1000000000
        w_sum=int(w_sum*1000000000); w_sum/=1000000000

        for i in range(2, nrow+1):
            x=input_sh.cell(i,2).value - u_sum; x=int(x*1000000000); x/=1000000000; Ud.append(x)
            y=input_sh.cell(i,3).value - v_sum; y=int(y*1000000000); y/=1000000000; Vd.append(y)
            z=input_sh.cell(i,4).value - w_sum; z=int(z*1000000000); z/=1000000000; Wd.append(z)
            o=get_octant(x,y,z); Ov.append(o)

        #  xlsx output file
        wb_output=Workbook()
        output_sh=wb_output.active

        # printing  initial values, average values, preprocessed values
        output_sh.cell(1,1, value="Time")
        output_sh.cell(1,2, value="U")
        output_sh.cell(1,3, value="V")
        output_sh.cell(1,4, value="W")
        for i in range(2, nrow+1):
            output_sh.cell(i,1, value=input_sh.cell(i,1).value)
            output_sh.cell(i,2, value=input_sh.cell(i,2).value)
            output_sh.cell(i,3, value=input_sh.cell(i,3).value)
            output_sh.cell(i,4, value=input_sh.cell(i,4).value)


        output_sh.cell(1,5, value='U Avg'); output_sh.cell(2,5, value=u_sum)
        output_sh.cell(1,6, value='V Avg'); output_sh.cell(2,6, value=v_sum)
        output_sh.cell(1,7, value='W Avg'); output_sh.cell(2,7, value=w_sum)

        output_sh.cell(1,8, value="U'=U - U avg")
        output_sh.cell(1,9, value="V'=V - V avg")
        output_sh.cell(1,10, value="W'=W - W avg")
        output_sh.cell(1,11, value="Octact")
        j=0
        for i in range(2,nrow+1):
            output_sh.cell(i,8, value=Ud[j])
            output_sh.cell(i,9, value=Vd[j])
            output_sh.cell(i,10, value=Wd[j])
            output_sh.cell(i,11, value=Ov[j])
            j+=1
        
        # longest subsequence  count
        LSC={1:[0,0], -1:[0,0], 2:[0,0], -2:[0,0], 3:[0,0], -3:[0,0], 4:[0,0], -4:[0,0]}

        # temp octant value list
        # terminator is last element
        tov=Ov.copy(); tov.append(0)

        # cuurent octant and its length
        co=Ov[0]; col=1
        LSC[co][0]=1; LSC[co][1]=1

        for i in range(1,n+1):
            if(tov[i]==co):
                col+=1
            else:
                if(col>LSC[co][0]):
                    LSC[co][0]=col
                    LSC[co][1]=1
                elif(col==LSC[co][0]):
                    LSC[co][1]+=1
                co=tov[i]
                col=1

